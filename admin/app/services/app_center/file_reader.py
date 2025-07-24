import json
import os
import fitz
from app.models.resource_center.resource_model import ResourceObjectMeta
from app.app import db
from app.utils.oss.oss_client import generate_download_url
from app.services.resource_center.resource_object_service import set_resource_icon
from app.app import app
from openpyxl import load_workbook
import xlrd
import platform
import subprocess
import stat
import math


def file_reader_node_execute(task_params, task_record, global_params):
    """
    根据任务参数和任务记录执行文档读取节点。并将其进行内容提取至指定格式
        src-format:
        transform-engine:
        tgt-format:
            文本：markdown，html，text
            图片：jpg，png，webp
    :param task_params:
    :param task_record:
    :param global_params:
    :return:
    """
    input_resources = task_params.get("input_resources", [])
    src_format = task_record.workflow_node_file_reader_config.get("src_format", "pdf")
    transform_engine = task_record.workflow_node_file_reader_config.get("engine", "PyMuPDF")
    tgt_format = task_record.workflow_node_file_reader_config.get("tgt_format", "jpg")
    output_resources = []
    input_resources = ResourceObjectMeta.query.filter(
        ResourceObjectMeta.id.in_(input_resources)
    ).all()
    app.logger.info("开始执行文件读取节点，输入资源：{}, 源格式：{}, 转换引擎：{}, 目标格式：{}".format(
        input_resources,
        src_format,
        transform_engine,
        tgt_format
    ))
    for resource in input_resources:
        if src_format == "pdf" and resource.resource_format == "pdf":
            if transform_engine == "PyMuPDF":
                output_resources.extend(pyMuPDFReader(resource, task_params))
        elif src_format in ("docx", "doc"):
            if transform_engine == "python-docx":
                pass
            elif transform_engine == "pandoc":
                output_resources.extend(pandocReader(resource, task_params))
        elif src_format in ('xlsx', 'xls'):
            if transform_engine == "openpyxl":
                output_resources.extend(execlReader(resource, task_params))
        elif src_format == "未知":
            pass
    node_results = {"output_resources": [resource for resource in output_resources]}
    task_record.task_result = json.dumps(node_results)
    db.session.add(task_record)
    db.session.commit()
    return task_record


def pyMuPDFReader(resource, task_params):
    """
    使用PyMuPDF读取PDF文件并转换为指定格式。
    :return:
    """
    pdf = fitz.open(resource.resource_path)
    dpi = task_params.get("dpi", 300)
    tgt_format = task_params.get("tgt_format", "png")
    max_pixels = task_params.get("max_pixels", 12_000_000)
    max_size_mb = task_params.get("max_size_mb", 10)
    output_resources = []
    for page_num in range(len(pdf)):
        page = pdf.load_page(page_num)
        # 获取原始页面尺寸(以点为单位)
        rect = page.rect
        width_pt = rect.width
        height_pt = rect.height
        # 计算当前DPI下的像素数
        width_px = int(width_pt * dpi / 72)
        height_px = int(height_pt * dpi / 72)
        total_pixels = width_px * height_px
        # 如果像素数超过限制，调整DPI
        if total_pixels > max_pixels:
            # 计算允许的最大DPI
            scale_factor = math.sqrt(max_pixels / total_pixels)
            adjusted_dpi = int(dpi * scale_factor)
        else:
            adjusted_dpi = dpi
        # 生成图像
        pix = page.get_pixmap(dpi=adjusted_dpi)
        target_new_name = resource.resource_name + f"_{page_num + 1}.{tgt_format}"
        target_new_path = resource.resource_path + f"_{page_num + 1}.{tgt_format}"
        pix.save(target_new_path, tgt_format)
        new_resource = ResourceObjectMeta(
            resource_parent_id=resource.id,
            user_id=resource.user_id,
            resource_name=target_new_name,
            resource_path=target_new_path,
            resource_type="image",
            resource_icon=set_resource_icon({
                "resource_type": "image",
                "resource_format": tgt_format
            }),
            resource_format=tgt_format,
            resource_size_in_MB=round(os.path.getsize(target_new_path) / (1024 * 1024), 2),
            resource_source="app_center",
            resource_download_url=generate_download_url(
                'app_center', target_new_path, suffix=tgt_format
            ).json.get("result", ""),
        )
        db.session.add(new_resource)
        db.session.flush()
        output_resources.append(
            {
                "id": new_resource.id,
                "format": new_resource.resource_format,
                "url": new_resource.resource_download_url,
                "content": new_resource.resource_download_url,
            }
        )
    db.session.commit()
    return output_resources


def execlReader(resource, task_params):
    """
    使用Excel读取器读取Excel文件并转换为指定格式。
    :return:
    """
    tgt_format = task_params.get("tgt_format", "text")
    output_resources = []
    if tgt_format == 'text':
        if resource.resource_format == 'xlsx':
            content = excel_to_strings_openpyxl(resource.resource_path)
        elif resource.resource_format == 'xls':
            content = excel_to_strings_xlrd(resource.resource_path)
        else:
            content = ''
        target_new_name = resource.resource_name + ".txt"
        target_new_path = resource.resource_path + ".txt"
        with open(target_new_path, 'w', encoding='utf-8') as f:
            f.write(content)
        new_resource = ResourceObjectMeta(
            resource_parent_id=resource.id,
            user_id=resource.user_id,
            resource_name=target_new_name,
            resource_path=target_new_path,
            resource_type="text",
            resource_icon=set_resource_icon({
                "resource_type": "text",
                "resource_format": "txt"
            }),
            resource_format="txt",
            resource_size_in_MB=round(os.path.getsize(resource.resource_path) / (1024 * 1024), 2),
            resource_source="app_center",
            resource_download_url=generate_download_url(
                'app_center', target_new_path, suffix='txt'
            ).json.get("result", ""),
        )
        db.session.add(new_resource)
        db.session.flush()
        output_resources.append({
            "id": new_resource.id,
            "format": new_resource.resource_format,
            "url": new_resource.resource_download_url,
            "content": content,
        })
        db.session.commit()
    return output_resources


def excel_to_strings_openpyxl(file_path):
    wb = load_workbook(file_path, rich_text=True)
    result = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        content = []
        # 获取表头
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else ' ')

        content.append("\t".join(headers))

        # 获取数据行
        for row in sheet.iter_rows(min_row=2):
            row_values = []
            for cell in row:
                value = str(cell.value) if cell.value is not None else ' '
                row_values.append(value)
            content.append("\t".join(row_values))
        result[sheet_name] = "\n".join(content)
    content = ''
    for sheet_name, sheet_content in result.items():
        content += f"Sheet: {sheet_name}\n{sheet_content}\n\n"
    return content


def excel_to_strings_xlrd(file_path):
    workbook = xlrd.open_workbook(file_path)
    result = {}

    for sheet in workbook.sheets():
        content = []
        # 处理表头
        headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        content.append("\t".join(headers))

        # 处理数据行
        for row in range(1, sheet.nrows):
            row_values = []
            for col in range(sheet.ncols):
                value = sheet.cell_value(row, col)
                row_values.append(str(value) if value != '' else ' ')
            content.append("\t".join(row_values))

        result[sheet.name] = "\n".join(content)
    content = ''
    for sheet_name, sheet_content in result.items():
        content += f"Sheet: {sheet_name}\n{sheet_content}\n\n"
    return content


def pandocReader(resource, task_params):
    """
    使用Pandoc读取文档并转换为指定格式。
        来源格式支持docx
        目标格式支持html, markdown, text等
    :return:
    """
    # 这里可以使用subprocess调用pandoc命令行工具进行转换
    tgt_format = task_params.get("tgt_format", "html")
    output_resources = []
    base_path = pickPandocLib()
    target_new_name = resource.resource_name + f".{tgt_format}"
    target_new_path = resource.resource_path + f".{tgt_format}"
    to_format = {
        "html": "html",
        "markdown": "markdown",
        "text": "plain",
    }.get(tgt_format, "plain")
    try:
        subprocess.run([
            base_path,
            resource.resource_path,
            '-o', target_new_path,
            '--from', resource.resource_format,
            '--to', to_format,
            '--embed-resources',
            '--standalone',
            '--preserve-tabs',
            '--wrap=preserve',
            '--variable=mainfont:"SimSun"',
            '--mathml',
        ], check=True)
    except subprocess.CalledProcessError as e:
        return f"转换失败: {str(e)}"
    new_resource = ResourceObjectMeta(
        resource_parent_id=resource.id,
        user_id=resource.user_id,
        resource_name=target_new_name,
        resource_path=target_new_path,
        resource_type="text",
        resource_icon=set_resource_icon({
            "resource_type": "document",
            "resource_format": tgt_format
        }),
        resource_format=tgt_format,
        resource_size_in_MB=round(os.path.getsize(target_new_path) / (1024 * 1024), 2),
        resource_source="app_center",
        resource_download_url=generate_download_url(
            'app_center', target_new_path, suffix=tgt_format
        ).json.get("result", ""),
    )
    db.session.add(new_resource)
    db.session.flush()
    with open(target_new_path, 'r', encoding='utf-8') as f:
        content = f.read()
    output_resources.append(
        {
            "id": new_resource.id,
            "format": new_resource.resource_format,
            "url": new_resource.resource_download_url,
            "content": content,
        }
    )
    db.session.commit()
    return output_resources


def pickPandocLib():
    """
    根据系统环境选择pandoc
    :return:
    """
    system = platform.system()
    machine = platform.machine()
    # 确定基本目录结构
    # 构建版本映射表
    version_map = {
        ('darwin', 'arm64'): 'pandoc-3.7.0.2-arm64-macOS',
        ('darwin', 'x86_64'): 'pandoc-3.7.0.2-x86_64-macOS',
        ('linux', 'x86_64'): 'pandoc-3.7.0.2-linux-amd64',
        ('linux', 'amd64'): 'pandoc-3.7.0.2-linux-amd64',
        ('linux', 'arm64'): 'pandoc-3.7.0.2-linux-arm64',
        ('windows', 'amd64'): 'pandoc-3.7.0.2-windows-x86_64',
        ('windows', 'x86_64'): 'pandoc-3.7.0.2-windows-x86_64'
    }
    target_version = version_map.get((system.lower(), machine.lower()))
    if not target_version:
        raise Exception("Unsupported platform or architecture for Pandoc.")
    pandoc_name = "pandoc" if system.lower() != 'windows' else "pandoc.exe"
    base_path = os.path.join(
        app.config['base_dir'], 'libs', 'pandoc', target_version, 'bin', pandoc_name
    )
    if not os.path.exists(base_path):
        raise Exception(f"Pandoc binary not found at {base_path}.")
    # 如果不是Windows系统，检查并设置执行权限
    if system != 'windows':
        # 获取当前文件权限
        current_mode = os.stat(base_path).st_mode
        # 检查是否有执行权限
        if not (current_mode & stat.S_IXUSR):
            try:
                # 添加用户执行权限
                os.chmod(base_path, current_mode | stat.S_IXUSR | stat.S_IXGRP | stat.S_IXOTH)
            except Exception as e:
                raise Exception(f"Failed to set execute permission for {base_path}: {str(e)}")
    return base_path


def tikaReader():
    """
    使用Apache Tika读取文档并转换为指定格式。
    :return:
    """
    # 这里可以使用tika-python库进行文档解析
    pass


def pyDocxReader():
    """
    使用python-docx读取Word文档并转换为指定格式。
    :return:
    """
    # 这里可以使用python-docx库进行Word文档解析
    pass


def textReader():
    """
    使用文本读取器读取纯文本文件并转换为指定格式。
    :return:
    """