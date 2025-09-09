import openpyxl
import pandas as pd
from datetime import datetime, timedelta
def trans_datetime(input_date_str):
    """
    # 202210 # 转为 2022-10-01 和 2022-10-31 两个值
    """
    # print(str(input_date_str))
    start_date = datetime.strptime(str(input_date_str), '%Y%m')
    # 开始时间
    start_time = start_date.strftime('%Y-%m-%d')

    # 获取该月的最后一天
    next_month = start_date.replace(day=28) + timedelta(days=4)
    end_date = next_month - timedelta(days=next_month.day)
    # 结束时间
    end_time = end_date.strftime('%Y-%m-%d')
    return start_time, end_time

data_dir = 'zy_data'
sheet_name = 'Sheet1'
zy_data = [
    {
        "name": "大户普通户投放-进货条数202210-202212数据.xlsx",
        "begin_row": 3,
        "begin_col": 4,
    }
]

all_data = [
    ["start_time", "end_time",  "product_level", "product_specification", 	"city",
     "customer_level", "avg_purchase", "avg_purchase_all"]
]



for item in zy_data:
    wb = openpyxl.load_workbook(f'{data_dir}/{item["name"]}')
    begin_row = item["begin_row"]
    begin_col = item["begin_col"]
    sheet = wb[sheet_name]
    start_col = begin_col
    # 每一行，每一列进行遍历，生成数据
    while True:
        product_level = sheet.cell(row=begin_row, column=1).value
        product_specification = sheet.cell(row=begin_row, column=3).value
        city = sheet.cell(row=begin_row, column=10).value
        record_time = sheet.cell(row=begin_row, column=11).value
        try:
            start_time, end_time = trans_datetime(record_time)
        except Exception as e:
            print(e)
            continue
        for col_idx in range(0, 3):
            customer_level = sheet.cell(row=2, column=begin_col + col_idx).value
            avg_purchase = sheet.cell(row=begin_row, column=begin_col + col_idx).value
            try:
                avg_purchase = float(avg_purchase)
            except Exception:
                avg_purchase = None
            avg_purchase_all = sheet.cell(row=begin_row, column=begin_col + col_idx + 3).value
            try:
                avg_purchase_all = float(avg_purchase_all)
            except Exception:
                avg_purchase_all = None
            all_data.append([start_time, end_time,
                             product_level, product_specification, city,
                             customer_level, avg_purchase, avg_purchase_all])
        begin_row += 1
        if not sheet.cell(row=begin_row, column=1).value:
            break
    # 关闭工作簿
    wb.close()
print(all_data)
df = pd.DataFrame(all_data[1:], columns=all_data[0])
print(df)
df.to_excel("zy_purchase_data.xlsx", index=False, header=True)
