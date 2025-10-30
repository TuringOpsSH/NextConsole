import pandas as pd
from app.services.configure_center.response_utils import next_console_response
from app.app import db
from app.utils.oss.oss_client import generate_new_path, generate_download_url
from app.models.resource_center.resource_model import ResourceObjectMeta
import os
import markdown
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import re


def md_table_to_excel_service(data):
    """
    将markdown表格转换为excel文件
    """
    md_table_text = data.get('md_table_text', '')
    user_id = data.get('user_id')
    filename = data.get('filename', 'md_table.xlsx')
    if not md_table_text:
        return next_console_response(error_status=True, error_message="md_table_text参数不能为空")

    try:
        html_content = markdown.markdown(md_table_text, extensions=['tables'])
        soup = BeautifulSoup(html_content, 'html.parser')
        # 查找所有表格
        tables = soup.find_all('table')
        if not tables:
            return next_console_response(error_status=True, error_message="未找到任何表格")

        new_resource_path = generate_new_path(
            module_name='app_center',
            user_id=user_id,
            suffix='xlsx'
        ).json.get("result")
        # 创建 Excel 工作簿
        wb = Workbook()
        # 删除默认创建的工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # 处理每个找到的表格
        for i, table in enumerate(tables):
            # 提取表格数据
            table_data = []
            for row in table.find_all('tr'):
                row_data = []
                for cell in row.find_all(['th', 'td']):
                    # 提取单元格文本内容，去除多余空白
                    cell_text = cell.get_text(strip=True)
                    row_data.append(cell_text)

                if row_data:  # 避免空行
                    table_data.append(row_data)

            if not table_data:
                continue

            # 生成工作表名称
            sheet_name = f"Table_{i + 1}"

            # 尝试使用表格前面的标题作为工作表名称
            prev_element = table.find_previous_sibling(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p'])
            if prev_element and prev_element.get_text(strip=True):
                title_text = prev_element.get_text(strip=True)
                # 清理标题文本，使其适合作为工作表名称
                clean_title = re.sub(r'[\\/*?:[\]]', '', title_text[:25])
                if clean_title:
                    sheet_name = clean_title

            # 确保工作表名称唯一且不超过31个字符
            base_sheet_name = sheet_name[:31]
            counter = 1
            while base_sheet_name in wb.sheetnames:
                base_sheet_name = f"{sheet_name[:28]}_{counter}"
                counter += 1

            # 创建工作表
            ws = wb.create_sheet(title=base_sheet_name)

            # 写入数据
            for row_idx, row_data in enumerate(table_data):
                for col_idx, cell_value in enumerate(row_data):
                    # 尝试将看起来像数字的值转换为数字
                    try:
                        if cell_value.replace('.', '', 1).isdigit():
                            cell_value = float(cell_value) if '.' in cell_value else int(cell_value)
                    except:
                        pass

                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

            # 设置表头样式（第一行）
            if table_data:
                for col_idx in range(1, len(table_data[0]) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

        # 保存 Excel 文件
        wb.save(new_resource_path)
    except Exception as e:
        return next_console_response(error_status=True, error_message=f"处理Markdown表格失败: {str(e)}")

    try:
        # # 生成下载链接
        resource_show_url = generate_download_url(
            module_name="app_center",
            file_path=new_resource_path,
            suffix='xlsx',
        ).json.get("result")
        new_resource = ResourceObjectMeta(
            user_id=user_id,
            resource_name=filename,
            resource_type="document",
            resource_format='xlsx',
            resource_size_in_MB=os.path.getsize(new_resource_path) / 1024 / 1024,
            resource_path=new_resource_path,
            resource_source_url=resource_show_url,
            resource_show_url=resource_show_url,
            resource_status="正常",
            resource_source='app_center'
        )
        db.session.add(new_resource)
        db.session.commit()
    except Exception as e:
        return next_console_response(error_status=True, error_message=f"生成Excel文件失败: {str(e)}")
    return next_console_response(result={
        "id": new_resource.id,
        "name": new_resource.resource_name,
        "url": new_resource.resource_show_url
    })


def extract_cell_content(cell_node):
    """
    提取单元格的内容，处理内联标记（如粗体、斜体等）
    """
    if not hasattr(cell_node, 'children') or not cell_node.children:
        return ""

    content_parts = []

    # 递归提取文本内容
    def extract_text(node):
        if node.type == 'text':
            return node.content
        elif node.type in ['strong', 'em', 'code_inline']:
            # 对于这些内联标记，我们提取其文本内容
            return ''.join(extract_text(child) for child in node.children)
        elif hasattr(node, 'children'):
            return ''.join(extract_text(child) for child in node.children)
        return ""

    return extract_text(cell_node)