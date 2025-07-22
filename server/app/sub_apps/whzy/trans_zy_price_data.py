import openpyxl
import pandas as pd
data_dir = 'zy_data'
sheet_name = '行情价格'
zy_data = [
    {
        "name": "行情价格周监测报告 2022年1024-1030 修改版.xlsx",
        "begin_row": 5,
        "begin_col": 12,
        "start_time": "2022-10-24",
        "end_time": "2022-10-30",
    },
    {
        "name": "1_行情价格周监测报告 2022年1031-1106.xlsx",
        "begin_row": 5,
        "begin_col": 12,
        "start_time": "2022-10-31",
        "end_time": "2022-11-06",
    },
    {
        "name": "行情价格周监测报告 2022年1107-1113.xlsx",
        "begin_row": 5,
        "begin_col": 12,
        "start_time": "2022-11-07",
        "end_time": "2022-11-13",
    },
    {
        "name": "行情价格周监测报告 2022年1114-1120.xlsx",
        "begin_row": 5,
        "begin_col": 20,
        "start_time": "2022-11-14",
        "end_time": "2022-11-20",
    },
    {
        "name": "第七期周报 2022年1205-1211.xlsx",
        "begin_row": 5,
        "begin_col": 20,
        "start_time": "2022-12-05",
        "end_time": "2022-12-11",
    },
    {
        "name": "第八期周报 2022年1212-1218（中烟未下订单）.xlsx",
        "begin_row": 5,
        "begin_col": 20,
        "start_time": "2022-12-12",
        "end_time": "2022-12-18",
    },
]

all_data = [
    ["start_time", "end_time",
     "product_name", "product_level", "product_specification",
     "wholesale_price",	"recommended_retail_price",	"city",
     "purchase_price", "sale_price"]
]
for item in zy_data:
    wb = openpyxl.load_workbook(f'{data_dir}/{item["name"]}')
    start_time = item["start_time"]
    end_time = item["end_time"]
    begin_row = item["begin_row"]
    begin_col = item["begin_col"]
    sheet = wb[sheet_name]
    start_col = begin_col
    product_name = sheet.cell(row=begin_row, column=1).value
    product_level = sheet.cell(row=begin_row, column=begin_col+1).value

    # 每一行，每一列进行遍历，生成数据
    while True:
        for col_idx in range(0, 16):
            purchase_price = sheet.cell(row=begin_row, column=begin_col + col_idx).value
            sale_price = sheet.cell(row=begin_row, column=begin_col + col_idx + 16).value
            try:
                purchase_price = float(purchase_price)
            except:
                purchase_price = None
            try:
                sale_price = float(sale_price)
            except:
                sale_price = None
            sub_product_name = sheet.cell(row=begin_row, column=1).value
            if sub_product_name:
                product_name = sub_product_name
            sub_product_level = sheet.cell(row=begin_row, column=2).value
            if sub_product_level:
                product_level = sub_product_level
            product_specification = sheet.cell(row=begin_row, column=4).value
            wholesale_price = sheet.cell(row=begin_row, column=5).value
            recommended_retail_price = sheet.cell(row=begin_row, column=6).value
            city = sheet.cell(row=4, column=begin_col + col_idx).value
            all_data.append([start_time, end_time,
                             product_name, product_level, product_specification,
                             wholesale_price, recommended_retail_price, city,
                             purchase_price, sale_price])
        begin_row += 1
        if not sheet.cell(row=begin_row, column=begin_col).value:
            break
    # 关闭工作簿
    wb.close()
print(all_data)
df = pd.DataFrame(all_data[1:], columns=all_data[0])
print(df)
df.to_excel("zy_price_data.xlsx", index=False, header=True)
